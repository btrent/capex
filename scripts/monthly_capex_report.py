"""Monthly CapEx report generator for software capitalization (ASC 350-40).

Pulls resolved Jira issues for a given month, joins with engineer cost data, and
emits an XLSX report broken down by project, engineer, and (engineer x project)
detail. Optionally uploads to Google Drive and emails a summary to Finance.

Usage:
    python monthly_capex_report.py --month 2026-03 --config config.yaml --out report.xlsx

Environment (required unless --dry-run):
    JIRA_EMAIL, JIRA_API_TOKEN
    SMTP_USER, SMTP_PASS (for email)
    GOOGLE_SERVICE_ACCOUNT_JSON (path or inline JSON for Drive upload)
"""
from __future__ import annotations

import argparse
import calendar
import csv
import hashlib
import json
import logging
import os
import smtplib
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Iterable

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_VERSION = "1.0.0"
FALLBACK_STORY_POINTS = 1.0
DEFAULT_WARNING_THRESHOLD = 25

log = logging.getLogger("capex")


# ---------------------------------------------------------------------------
# Domain types
# ---------------------------------------------------------------------------
@dataclass
class Engineer:
    account_id: str
    name: str
    monthly_fully_loaded_cost: float
    start_date: date | None = None
    end_date: date | None = None

    def prorated_cost(self, month_start: date, month_end: date) -> float:
        """Prorate monthly cost if the engineer joined/left mid-month."""
        days_in_month = (month_end - month_start).days + 1
        active_start = max(self.start_date or month_start, month_start)
        active_end = min(self.end_date or month_end, month_end)
        if active_end < active_start:
            return 0.0
        active_days = (active_end - active_start).days + 1
        if active_days >= days_in_month:
            return self.monthly_fully_loaded_cost
        return self.monthly_fully_loaded_cost * (active_days / days_in_month)


@dataclass
class Epic:
    key: str
    summary: str
    capex_eligible: str  # "Yes" / "No" / "TBD"
    capex_stage: str | None  # "Application Development" / ...
    project_code: str | None
    placed_in_service: date | None

    @property
    def is_capex(self) -> bool:
        return (
            self.capex_eligible == "Yes"
            and self.capex_stage == "Application Development"
            and bool(self.project_code)
        )


@dataclass
class Ticket:
    key: str
    assignee_account_id: str | None
    assignee_name: str | None
    story_points: float | None
    parent_epic_key: str | None
    resolution_date: date | None


@dataclass
class Warning:
    category: str
    message: str
    ticket_key: str = ""


@dataclass
class Allocation:
    engineer: Engineer
    monthly_cost: float
    capex_points: float
    total_points: float
    capex_cost: float
    non_capex_cost: float
    project_points: dict[str, float] = field(default_factory=dict)

    @property
    def capex_ratio(self) -> float:
        return self.capex_points / self.total_points if self.total_points else 0.0


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
@dataclass
class Config:
    jira_base_url: str
    custom_fields: dict[str, str]
    jql_denominator: str
    jql_capex_filter: str
    engineer_costs_path: str
    projects_path: str
    drive_folder_id: str
    email_from: str
    email_to: list[str]
    smtp_host: str
    smtp_port: int
    warning_count_max: int
    raw: dict[str, Any]

    @classmethod
    def load(cls, path: str) -> "Config":
        with open(path) as f:
            raw = yaml.safe_load(f)
        jira = raw["jira"]
        delivery = raw.get("delivery", {})
        email_cfg = delivery.get("email", {})
        thresholds = raw.get("thresholds", {})
        return cls(
            jira_base_url=jira["base_url"],
            custom_fields=jira["custom_fields"],
            jql_denominator=jira.get("jql_denominator", "resolutiondate >= {start} AND resolutiondate < {end}"),
            jql_capex_filter=jira.get("jql_capex_filter", ""),
            engineer_costs_path=raw["paths"]["engineer_costs"],
            projects_path=raw["paths"]["projects"],
            drive_folder_id=delivery.get("drive_folder_id", ""),
            email_from=email_cfg.get("from", ""),
            email_to=list(email_cfg.get("to", [])),
            smtp_host=email_cfg.get("smtp_host", "smtp.gmail.com"),
            smtp_port=int(email_cfg.get("smtp_port", 587)),
            warning_count_max=int(thresholds.get("warning_count_max", DEFAULT_WARNING_THRESHOLD)),
            raw=raw,
        )

    def hash(self) -> str:
        return hashlib.sha256(json.dumps(self.raw, sort_keys=True, default=str).encode()).hexdigest()[:12]


# ---------------------------------------------------------------------------
# Jira client (thin wrapper; network isolation boundary)
# ---------------------------------------------------------------------------
class JiraClient:
    """Thin wrapper around the `jira` SDK. All network I/O lives here."""

    def __init__(self, base_url: str, email: str, api_token: str):
        from jira import JIRA  # local import so tests don't require the SDK

        self._client = JIRA(server=base_url, basic_auth=(email, api_token))

    def search(self, jql: str, fields: list[str]) -> list[dict]:
        """Return raw issue dicts for all matching issues (paginated)."""
        all_issues: list[dict] = []
        start_at = 0
        page_size = 100
        while True:
            page = self._client.search_issues(
                jql, startAt=start_at, maxResults=page_size, fields=",".join(fields), expand="names"
            )
            for issue in page:
                all_issues.append(issue.raw)
            if len(page) < page_size:
                break
            start_at += page_size
        return all_issues

    def get_epic(self, key: str, fields: list[str]) -> dict:
        issue = self._client.issue(key, fields=",".join(fields))
        return issue.raw


# ---------------------------------------------------------------------------
# Month parsing
# ---------------------------------------------------------------------------
def month_bounds(month_str: str) -> tuple[date, date]:
    """Given '2026-03', return (2026-03-01, 2026-03-31)."""
    year, mo = (int(x) for x in month_str.split("-"))
    start = date(year, mo, 1)
    end = date(year, mo, calendar.monthrange(year, mo)[1])
    return start, end


def default_prior_month() -> str:
    today = date.today()
    first_of_this_month = today.replace(day=1)
    last_of_prior = first_of_this_month - timedelta(days=1)
    return last_of_prior.strftime("%Y-%m")


# ---------------------------------------------------------------------------
# Parsing Jira issues
# ---------------------------------------------------------------------------
def parse_epic(issue: dict, custom_fields: dict[str, str]) -> Epic:
    fields = issue["fields"]
    eligible = _field_value(fields.get(custom_fields["capex_eligible"]))
    stage = _field_value(fields.get(custom_fields["capex_stage"]))
    project_code = _field_value(fields.get(custom_fields["capex_project_code"]))
    pis = _parse_date(fields.get(custom_fields["placed_in_service_date"]))
    return Epic(
        key=issue["key"],
        summary=fields.get("summary", ""),
        capex_eligible=eligible or "TBD",
        capex_stage=stage,
        project_code=project_code,
        placed_in_service=pis,
    )


def parse_ticket(issue: dict, custom_fields: dict[str, str]) -> Ticket:
    fields = issue["fields"]
    assignee = fields.get("assignee") or {}
    parent = fields.get("parent") or {}
    sp_raw = fields.get(custom_fields["story_points"])
    return Ticket(
        key=issue["key"],
        assignee_account_id=assignee.get("accountId"),
        assignee_name=assignee.get("displayName"),
        story_points=float(sp_raw) if sp_raw not in (None, "") else None,
        parent_epic_key=parent.get("key"),
        resolution_date=_parse_date(fields.get("resolutiondate")),
    )


def _field_value(raw: Any) -> str | None:
    """Select fields come back as {'value': 'Yes'}; text fields as str."""
    if raw is None:
        return None
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name")
    return str(raw)


def _parse_date(raw: Any) -> date | None:
    if not raw:
        return None
    if isinstance(raw, date):
        return raw
    s = str(raw)
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            parsed = datetime.strptime(s[: len(fmt) + 6] if "%z" in fmt else s[: len(fmt)], fmt)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except ValueError:
            continue
    # Fallback: first 10 chars look like YYYY-MM-DD
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Engineer cost table
# ---------------------------------------------------------------------------
def load_engineers(csv_path: str) -> dict[str, Engineer]:
    engineers: dict[str, Engineer] = {}
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            engineers[row["jira_account_id"]] = Engineer(
                account_id=row["jira_account_id"],
                name=row["name"],
                monthly_fully_loaded_cost=float(row["monthly_fully_loaded_cost"]),
                start_date=_parse_date(row.get("start_date")),
                end_date=_parse_date(row.get("end_date")),
            )
    return engineers


def load_projects(yaml_path: str) -> dict[str, dict]:
    if not os.path.exists(yaml_path):
        return {}
    with open(yaml_path) as f:
        data = yaml.safe_load(f) or {}
    return {p["code"]: p for p in data.get("projects", [])}


# ---------------------------------------------------------------------------
# Allocation math (pure, testable)
# ---------------------------------------------------------------------------
def allocate(
    tickets: list[Ticket],
    epics: dict[str, Epic],
    engineers: dict[str, Engineer],
    month_start: date,
    month_end: date,
) -> tuple[list[Allocation], list[Warning]]:
    """Compute per-engineer allocations and per-project breakdown.

    Returns allocations plus any warnings accumulated during processing.
    Engineers with zero resolved tickets are excluded (not divide-by-zero).
    Engineers missing from the cost CSV are skipped with a warning.
    """
    warnings: list[Warning] = []
    # engineer_id -> {"capex_points": x, "total_points": x, "project_points": {code: pts}}
    totals: dict[str, dict[str, Any]] = {}

    for t in tickets:
        if not t.assignee_account_id:
            continue
        points = t.story_points
        if points is None:
            points = FALLBACK_STORY_POINTS
            warnings.append(
                Warning("missing_story_points", f"Ticket {t.key} has no story points; using fallback=1", t.key)
            )
        bucket = totals.setdefault(
            t.assignee_account_id, {"capex_points": 0.0, "total_points": 0.0, "project_points": {}, "name": t.assignee_name}
        )
        bucket["total_points"] += points

        epic = epics.get(t.parent_epic_key) if t.parent_epic_key else None
        if epic and epic.is_capex:
            bucket["capex_points"] += points
            code = epic.project_code or "UNCODED"
            bucket["project_points"][code] = bucket["project_points"].get(code, 0.0) + points

    allocations: list[Allocation] = []
    for acct_id, bucket in totals.items():
        if bucket["total_points"] == 0:
            continue  # excluded entirely, never divide-by-zero
        engineer = engineers.get(acct_id)
        if engineer is None:
            warnings.append(
                Warning("missing_engineer", f"Assignee {acct_id} ({bucket['name']}) not in cost CSV; skipping")
            )
            continue
        monthly_cost = engineer.prorated_cost(month_start, month_end)
        ratio = bucket["capex_points"] / bucket["total_points"]
        capex_cost = monthly_cost * ratio
        allocations.append(
            Allocation(
                engineer=engineer,
                monthly_cost=monthly_cost,
                capex_points=bucket["capex_points"],
                total_points=bucket["total_points"],
                capex_cost=capex_cost,
                non_capex_cost=monthly_cost - capex_cost,
                project_points=dict(bucket["project_points"]),
            )
        )
    return allocations, warnings


def project_costs(allocations: list[Allocation]) -> dict[str, dict[str, float]]:
    """Map project_code -> {"cost": $, "points": pts, "contributors": n}."""
    projects: dict[str, dict[str, float]] = {}
    for alloc in allocations:
        if alloc.capex_points == 0:
            continue
        for code, points in alloc.project_points.items():
            share = alloc.capex_cost * (points / alloc.capex_points)
            entry = projects.setdefault(code, {"cost": 0.0, "points": 0.0, "contributors": 0})
            entry["cost"] += share
            entry["points"] += points
            entry["contributors"] += 1
    return projects


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


def _write_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _autosize(ws) -> None:
    for col in ws.columns:
        values = [str(c.value) if c.value is not None else "" for c in col]
        width = min(max(len(v) for v in values) + 2, 50) if values else 12
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def write_xlsx(
    path: str,
    month_str: str,
    allocations: list[Allocation],
    projects: dict[str, dict[str, float]],
    epics: dict[str, Epic],
    project_registry: dict[str, dict],
    warnings: list[Warning],
    config: Config,
    jql_used: str,
) -> None:
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    total_capex = sum(a.capex_cost for a in allocations)
    n_projects = len(projects)
    n_engineers = sum(1 for a in allocations if a.capex_cost > 0)
    n_epics = sum(1 for e in epics.values() if e.is_capex)
    ws.append(["Metric", "Value"])
    ws.append(["Month", month_str])
    ws.append(["Total Capitalized", round(total_capex, 2)])
    ws.append(["# Engineers Contributing", n_engineers])
    ws.append(["# Projects", n_projects])
    ws.append(["# Capex Epics", n_epics])
    ws.append([])
    ws.append(["Top 5 Projects by $"])
    ws.append(["Project Code", "Cost"])
    top5 = sorted(projects.items(), key=lambda kv: kv[1]["cost"], reverse=True)[:5]
    for code, data in top5:
        ws.append([code, round(data["cost"], 2)])
    for cell in ws["A"]:
        cell.font = Font(bold=True) if cell.row in (1, 8, 9) else cell.font

    # By Project
    ws = wb.create_sheet("By Project")
    _write_header(ws, ["Project Code", "Name", "Total $", "# Contributors", "# Points"])
    for code, data in sorted(projects.items(), key=lambda kv: kv[1]["cost"], reverse=True):
        name = project_registry.get(code, {}).get("name", "")
        ws.append([code, name, round(data["cost"], 2), int(data["contributors"]), round(data["points"], 2)])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.00"

    # By Engineer
    ws = wb.create_sheet("By Engineer")
    _write_header(ws, ["Name", "Monthly Loaded Cost", "CapEx Ratio %", "CapEx $", "Non-CapEx $"])
    for a in sorted(allocations, key=lambda x: x.capex_cost, reverse=True):
        ws.append([
            a.engineer.name,
            round(a.monthly_cost, 2),
            round(a.capex_ratio * 100, 2),
            round(a.capex_cost, 2),
            round(a.non_capex_cost, 2),
        ])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"

    # Detail (engineer x project)
    ws = wb.create_sheet("Detail")
    _write_header(ws, ["Engineer", "Project Code", "Points", "$"])
    for a in allocations:
        if a.capex_points == 0:
            continue
        for code, points in sorted(a.project_points.items()):
            share = a.capex_cost * (points / a.capex_points)
            ws.append([a.engineer.name, code, round(points, 2), round(share, 2)])

    # Warnings
    ws = wb.create_sheet("Warnings")
    _write_header(ws, ["Category", "Ticket", "Message"])
    for w in warnings:
        ws.append([w.category, w.ticket_key, w.message])

    # Metadata
    ws = wb.create_sheet("Metadata")
    ws.append(["Run Timestamp (UTC)", datetime.utcnow().isoformat(timespec="seconds")])
    ws.append(["Month Covered", month_str])
    ws.append(["JQL Used", jql_used])
    ws.append(["Config Hash", config.hash()])
    ws.append(["Script Version", SCRIPT_VERSION])
    ws.append(["Warning Count", len(warnings)])

    for sheet in wb.worksheets:
        _autosize(sheet)

    wb.save(path)


# ---------------------------------------------------------------------------
# Delivery
# ---------------------------------------------------------------------------
def upload_to_drive(path: str, folder_id: str) -> str | None:
    """Upload the XLSX to Drive; return file id. Requires GOOGLE_SERVICE_ACCOUNT_JSON."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
    except ImportError:
        log.warning("google-api-python-client not installed; skipping Drive upload")
        return None

    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not sa_json or not folder_id:
        log.warning("Drive upload skipped (missing service account or folder id)")
        return None
    info = json.loads(Path(sa_json).read_text()) if os.path.exists(sa_json) else json.loads(sa_json)
    creds = service_account.Credentials.from_service_account_info(info, scopes=["https://www.googleapis.com/auth/drive.file"])
    service = build("drive", "v3", credentials=creds, cache_discovery=False)
    meta = {"name": os.path.basename(path), "parents": [folder_id]}
    media = MediaFileUpload(path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    result = service.files().create(body=meta, media_body=media, fields="id,webViewLink").execute()
    log.info("Uploaded to Drive: %s", result.get("webViewLink"))
    return result.get("id")


def send_email(config: Config, subject: str, body: str) -> None:
    user = os.environ.get("SMTP_USER")
    pw = os.environ.get("SMTP_PASS")
    if not user or not pw or not config.email_to:
        log.warning("Email skipped (missing SMTP creds or recipients)")
        return
    msg = MIMEMultipart()
    msg["From"] = config.email_from or user
    msg["To"] = ", ".join(config.email_to)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    with smtplib.SMTP(config.smtp_host, config.smtp_port) as s:
        s.starttls()
        s.login(user, pw)
        s.send_message(msg)
    log.info("Email sent to %s", config.email_to)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def fetch_jira_data(
    client: JiraClient, config: Config, month_start: date, month_end: date
) -> tuple[list[Ticket], dict[str, Epic], str]:
    """Pull tickets + epics for the month. Returns (tickets, epics_by_key, jql_used)."""
    cf = config.custom_fields
    start_str = month_start.isoformat()
    end_str = (month_end + timedelta(days=1)).isoformat()
    jql = config.jql_denominator.format(start=f'"{start_str}"', end=f'"{end_str}"')
    fields = [
        "summary", "assignee", "resolutiondate", "parent", "issuetype", "status",
        cf["story_points"],
    ]
    log.info("JQL: %s", jql)
    raw_issues = client.search(jql, fields)
    if not raw_issues:
        raise RuntimeError("No issues returned from Jira — likely a misconfiguration")

    tickets = [parse_ticket(i, cf) for i in raw_issues]

    # Fetch parent epics in a separate pass — Jira Cloud rarely returns custom
    # fields on the parent object, so we re-fetch by key.
    epic_keys = {t.parent_epic_key for t in tickets if t.parent_epic_key}
    epic_fields = [
        "summary",
        cf["capex_eligible"], cf["capex_stage"], cf["capex_project_code"], cf["placed_in_service_date"],
    ]
    epics: dict[str, Epic] = {}
    for key in epic_keys:
        try:
            raw = client.get_epic(key, epic_fields)
            epics[key] = parse_epic(raw, cf)
        except Exception as exc:  # noqa: BLE001
            log.warning("Failed to fetch epic %s: %s", key, exc)
    return tickets, epics, jql


def run(args: argparse.Namespace) -> int:
    month_str = args.month or default_prior_month()
    month_start, month_end = month_bounds(month_str)
    config = Config.load(args.config)

    log.info("Running capex report for %s", month_str)

    email = os.environ.get("JIRA_EMAIL")
    token = os.environ.get("JIRA_API_TOKEN")
    if not email or not token:
        log.error("JIRA_EMAIL / JIRA_API_TOKEN not set")
        return 2

    try:
        client = JiraClient(config.jira_base_url, email, token)
    except Exception as exc:  # noqa: BLE001
        log.error("Jira auth failed: %s", exc)
        return 2

    try:
        tickets, epics, jql_used = fetch_jira_data(client, config, month_start, month_end)
    except RuntimeError as exc:
        log.error("%s", exc)
        return 3

    engineers = load_engineers(config.engineer_costs_path)
    project_registry = load_projects(config.projects_path)

    allocations, warnings = allocate(tickets, epics, engineers, month_start, month_end)
    projects = project_costs(allocations)

    write_xlsx(args.out, month_str, allocations, projects, epics, project_registry, warnings, config, jql_used)
    log.info("Wrote %s", args.out)

    total = sum(a.capex_cost for a in allocations)
    log.info("Total capitalized: $%.2f across %d engineers / %d projects", total, len(allocations), len(projects))

    if len(warnings) > config.warning_count_max:
        log.error("Warning count %d exceeds threshold %d", len(warnings), config.warning_count_max)
        return 4

    if not args.dry_run:
        upload_to_drive(args.out, config.drive_folder_id)
        body = (
            f"CapEx report for {month_str}\n\n"
            f"Total capitalized: ${total:,.2f}\n"
            f"Engineers: {len(allocations)} | Projects: {len(projects)} | Epics: {sum(1 for e in epics.values() if e.is_capex)}\n"
            f"Warnings: {len(warnings)}\n"
        )
        send_email(config, f"[CapEx] {month_str} report", body)
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--month", help="YYYY-MM (defaults to prior month)")
    parser.add_argument("--config", required=True, help="Path to config.yaml")
    parser.add_argument("--out", required=True, help="Output XLSX path")
    parser.add_argument("--dry-run", action="store_true", help="Skip Drive upload + email")
    parser.add_argument("--verbose", action="store_true", help="DEBUG logging")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    return run(args)


if __name__ == "__main__":
    sys.exit(main())
