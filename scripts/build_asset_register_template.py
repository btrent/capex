"""Build the accounting asset register XLSX template.

Generates /templates/capex_asset_register.xlsx with six tabs:
Instructions, Asset Register, Monthly Additions, Amortization Schedule,
Summary Dashboard, Journal Entry Helper.

Usage:
    python build_asset_register_template.py [--out PATH]
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DEFAULT_OUT = Path(__file__).resolve().parents[1] / "templates" / "capex_asset_register.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

MONEY_FMT = "#,##0.00"
DATE_FMT = "yyyy-mm-dd"


def _write_headers(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _autosize(ws, min_w: int = 12, max_w: int = 40) -> None:
    for col in ws.columns:
        values = [str(c.value) if c.value is not None else "" for c in col]
        width = min(max(len(v) for v in values) + 2, max_w) if values else min_w
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(width, min_w)


def _build_instructions(ws) -> None:
    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "CapEx Asset Register — Instructions"
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    lines = [
        "Purpose: Maintain the list of capitalized software projects, their amortization, and net book value.",
        "Cadence: Update monthly after the engineering CapEx report is delivered (1st of each month).",
        "",
        "Workflow each month:",
        "1. Open the latest monthly CapEx report XLSX (delivered via Drive + email).",
        "2. On the 'Monthly Additions' tab, append one row per project_code with the $ added this month.",
        "3. On the 'Asset Register' tab, update 'Total Capitalized' for each project (sum of all additions to date).",
        "4. Set 'In-Service Date' the FIRST month the project delivers production value. Amortization begins then.",
        "5. 'Monthly Amortization' and 'Net Book Value' are formulas — do not overwrite unless useful life changes.",
        "6. On 'Journal Entry Helper' copy the month's suggested JE lines to your GL (NetSuite, QBO, etc.).",
        "7. If a project is retired or impaired, set Status accordingly and zero out future amortization manually.",
        "",
        "Conventions:",
        "- Project Code must match the code used in the Jira CapEx Project Code field.",
        "- All dollar amounts in USD. All dates ISO format (yyyy-mm-dd).",
        "- Useful life defaults to 36 months (3 years straight-line). Finance can override per project.",
        "",
        "Questions? finance@yourcompany.com",
    ]
    for i, line in enumerate(lines, 2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = WRAP
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20


def _build_asset_register(ws) -> None:
    headers = [
        "Project Code", "Description", "In-Service Date", "Total Capitalized",
        "Useful Life (months)", "Monthly Amortization", "Accumulated Amortization",
        "Net Book Value", "Status",
    ]
    _write_headers(ws, headers)

    # Sample rows
    samples = [
        ["CAPEX-2026-PLATFORM-SSO", "SSO Platform Rollout", "2026-03-15", 185000.00, 36, None, 0.00, None, "Active"],
        ["CAPEX-2026-RESIDENT-APP", "Resident App v2 checkout flow", "2026-04-01", 92500.00, 36, None, 0.00, None, "Active"],
    ]
    for i, row in enumerate(samples, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        # Monthly Amortization = D/E
        ws.cell(row=i, column=6, value=f"=D{i}/E{i}")
        # Net Book Value = D - G
        ws.cell(row=i, column=8, value=f"=D{i}-G{i}")
        ws.cell(row=i, column=3).number_format = DATE_FMT
        for col in (4, 6, 7, 8):
            ws.cell(row=i, column=col).number_format = MONEY_FMT

    # Autofilter on the header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(samples) + 1}"
    _autosize(ws, max_w=30)


def _build_monthly_additions(ws) -> None:
    headers = ["Month", "Project Code", "Amount Added", "Source Report File", "Notes"]
    _write_headers(ws, headers)
    samples = [
        ["2026-03", "CAPEX-2026-PLATFORM-SSO", 125000.00, "capex-report-2026-03.xlsx", "Initial build"],
        ["2026-03", "CAPEX-2026-RESIDENT-APP", 55000.00, "capex-report-2026-03.xlsx", "Scaffolding + payment form"],
        ["2026-04", "CAPEX-2026-PLATFORM-SSO", 60000.00, "capex-report-2026-04.xlsx", "Finalize + ship"],
        ["2026-04", "CAPEX-2026-RESIDENT-APP", 37500.00, "capex-report-2026-04.xlsx", ""],
    ]
    for i, row in enumerate(samples, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        ws.cell(row=i, column=3).number_format = MONEY_FMT
    _autosize(ws, max_w=35)


def _build_amortization_schedule(ws) -> None:
    headers = ["Month #", "Period (yyyy-mm)", "Project Code", "Monthly Amortization", "Cumulative Amortization", "Remaining NBV"]
    _write_headers(ws, headers)
    # Sample 36-month schedule for the first register row (SSO).
    # Pulls Total Capitalized and Useful Life from 'Asset Register' row 2.
    # yyyy-mm period shown as literal starting 2026-03.
    for m in range(1, 37):
        r = m + 1
        # Simple year/month increment starting from March 2026
        month_idx = 2 + (m - 1)  # march=3 ... but +2 shifts; we'll build below
        year = 2026 + ((2 + m - 1) // 12)
        mo = ((2 + m - 1) % 12) + 1
        period = f"{year}-{mo:02d}"
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=period)
        ws.cell(row=r, column=3, value="CAPEX-2026-PLATFORM-SSO")
        # Monthly amortization = Asset Register D2 / E2
        ws.cell(row=r, column=4, value="='Asset Register'!D2/'Asset Register'!E2")
        # Cumulative = monthly * m
        ws.cell(row=r, column=5, value=f"=D{r}*A{r}")
        # Remaining NBV = Total Capitalized - Cumulative
        ws.cell(row=r, column=6, value=f"='Asset Register'!D2-E{r}")
        for col in (4, 5, 6):
            ws.cell(row=r, column=col).number_format = MONEY_FMT

    # Note row for accountant
    note_row = 39
    ws.cell(row=note_row, column=1, value="Note:")
    ws.cell(row=note_row, column=1).font = BOLD
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=6)
    ws.cell(
        row=note_row, column=2,
        value="Copy rows 2-37 and adjust formulas/project codes for each additional project in the Asset Register.",
    )
    _autosize(ws, max_w=35)


def _build_summary_dashboard(ws) -> None:
    ws["A1"] = "Summary Dashboard"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    rows = [
        ("Metric", "Value", ""),
        ("Total Capitalized (lifetime)", "=SUM('Asset Register'!D:D)", MONEY_FMT),
        ("Total Accumulated Amortization", "=SUM('Asset Register'!G:G)", MONEY_FMT),
        ("Total Net Book Value", "=SUM('Asset Register'!H:H)", MONEY_FMT),
        ("YTD Additions (current year — update range as needed)", "=SUMIF('Monthly Additions'!A:A,\">=2026-01\",'Monthly Additions'!C:C)", MONEY_FMT),
        ("", "", ""),
        ("Pivot hint:", "To convert into a pivot: select Asset Register and insert pivot.", ""),
    ]
    for i, (label, value, fmt) in enumerate(rows, 3):
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=value)
        if i == 3:
            a.font = HEADER_FONT
            a.fill = HEADER_FILL
            b.font = HEADER_FONT
            b.fill = HEADER_FILL
        if fmt:
            b.number_format = fmt
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 40


def _build_journal_entry_helper(ws) -> None:
    headers = ["Date", "Account #", "Account Name", "Dr", "Cr", "Memo"]
    _write_headers(ws, headers)
    rows = [
        # Monthly capitalization
        ["2026-03-31", "1700", "Capitalized Software (Asset)", 180000.00, None, "Mar 2026 capex per monthly report"],
        ["2026-03-31", "6100", "Salaries — Capitalized Offset", None, 180000.00, "Mar 2026 capex offset to engineering salaries"],
        # Monthly amortization
        ["2026-03-31", "7400", "Amortization Expense", 5139.00, None, "Mar 2026 amortization — CAPEX-2026-PLATFORM-SSO"],
        ["2026-03-31", "1701", "Accum. Amortization — Capitalized Software", None, 5139.00, "Mar 2026 amortization — CAPEX-2026-PLATFORM-SSO"],
        # Impairment
        ["2026-06-30", "7500", "Impairment Loss", 20000.00, None, "Impair CAPEX-2026-RESIDENT-APP — project descoped"],
        ["2026-06-30", "1700", "Capitalized Software (Asset)", None, 20000.00, "Impair CAPEX-2026-RESIDENT-APP — project descoped"],
        # Retirement
        ["2029-03-31", "1701", "Accum. Amortization — Capitalized Software", 185000.00, None, "Retire fully amortized CAPEX-2026-PLATFORM-SSO"],
        ["2029-03-31", "1700", "Capitalized Software (Asset)", None, 185000.00, "Retire fully amortized CAPEX-2026-PLATFORM-SSO"],
    ]
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
        ws.cell(row=i, column=1).number_format = DATE_FMT
        for col in (4, 5):
            ws.cell(row=i, column=col).number_format = MONEY_FMT
    _autosize(ws, max_w=45)


def build(out_path: Path) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    _build_instructions(ws)

    _build_asset_register(wb.create_sheet("Asset Register"))
    _build_monthly_additions(wb.create_sheet("Monthly Additions"))
    _build_amortization_schedule(wb.create_sheet("Amortization Schedule"))
    _build_summary_dashboard(wb.create_sheet("Summary Dashboard"))
    _build_journal_entry_helper(wb.create_sheet("Journal Entry Helper"))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="Output XLSX path")
    args = parser.parse_args()
    path = build(Path(args.out))
    print(f"Wrote {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
